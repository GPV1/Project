import openpyxl
from openpyxl.styles import Font
import pandas as pd
import numpy as np
from collections import defaultdict
import pprint

# Open workbook with Biovolume conversion data and Species order
wbBioV = openpyxl.load_workbook('C:/GIS_Course/EGM722/Project/Data_files/PEG_BVOL2019_PJ.xlsx')
wbSpOr = openpyxl.load_workbook('C:/GIS_Course/EGM722/Project/Data_files/Species_order_FG.xlsx')

# Check the sheet names of the workbook
print(wbBioV.sheetnames, wbSpOr.sheetnames)

sheetBio = wbBioV['Biovolume file']

# print the headers of the sheet
#for i in range(1, 34, 1):
    #print(i, sheetBio.cell(row=1, column=i).value)

# Choose from the headers and select the right column (biovolume)
print(sheetBio.cell(row=1, column=26).value)

# Prints all the data in the sheet
#for r in range(1,sheetBio.max_row+1 ,1):
    #for c in range(1,sheetBio.max_column+1,1):
        #print(r, sheetBio.cell(row=r, column=c).value)

# todo write a dict for the species groups
# todo define taxonomy groups of the species order list

dfBV = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/PEG_BVOL2019_PJ.xlsx',sheet_name='Biovolume file')
dfSpN = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/Species_order_GPV.xlsx',sheet_name='Sheet1')

pprint.pprint(dfBV.iterrows())

# Define variables for taxonomy groups
divisions = dfBV['Division']  #.unique()
Classes = dfBV['Class']  #.unique()
Order = dfBV['Order']   #.unique()
Genus = dfBV['Genus']  #.unique()
Species = dfBV['Species']  #.unique()

SpGe = dict(zip(Species, Genus))
SpOr = dict(zip(Species, Order))
SpCl = dict(zip(Species, Classes))
pprint.pprint(SpCl)
print('The species Gymnodiniales is from the class {}'.format(SpCl['Gymnodiniales']))

# define variable for species code (SC) and species name (SN)
SC = dfSpN['spec_code']
SN = dfSpN['spec_name']
SNs = dfSpN['spec_name_short']

# create dict with species code and species name, species code as key, species name and species name short as value
# todo find out how to make a nested dictionary to combine short name with spec name
spec_code = dict(zip(SC, SN))
spec_code_short = dict(zip(SC, SNs))

# tests to see dicts
#pprint.pprint(spec_code)
#pprint.pprint(spec_code_short)
#print(spec_code[138])
#print(spec_code_short[138])



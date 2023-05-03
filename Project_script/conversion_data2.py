import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import pandas as pd
import pprint

# import datafiles and sheetname. Necessary: database files PEG and Species order and input data
dfBiovolume = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/PEG_BVOL2019_PJ.xlsx', sheet_name='Biovolume file')
dfSpeciesOrder = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/Species_order_GPV.xlsx', sheet_name='Sheet1')
InputData = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/Data_2018_GPV.xlsx', sheet_name='Sheet1')

'''  preperation of database by forming dictionaries for taxonomy groups, creating averages of multiple biovolume
 values per species and usage of dictionaries in functions  '''

#  defines the columns in the database for the taxonomy groups
Divisions = dfBiovolume['Division']
Classes = dfBiovolume['Class']
Order = dfBiovolume['Order'].str.title()  # Order names in PEG database are in capitals
Genus = dfBiovolume['Genus']
Species = dfBiovolume['Species']
Species_other = dfSpeciesOrder['spec_name'].str.replace('_', ' ')  # Species names in SpeciesOrder database have _
Species_other_order = dfSpeciesOrder['order']
Species_other_group = dfSpeciesOrder['group']
All_Species = pd.concat([pd.Series(Species), pd.Series(Species_other)])

InputData['spec_name'] = InputData['spec_name'].str.replace('.', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('<', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('>', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('~', '')

'----------------------------------------------------------------------------------------------------------------------'
''' Overview of all dictonaries and their functions:

- 1. species_Biovolume_dict:                    Species from the PEG database with their corresponding biovolume values 
                                                (multiple values per species). 
                                                * {Keys: Species, values: biovolume}.
                                     
- 2. species_avg_Biovolume_dict:                Average biovolume for species from the PEG database. 
                                                * {Keys: Species, values: average biovolume per species}.

- 3. Order_Species_avg_Biovolume_dict:          The average biovolume per species is nested in the corresponding order. 
                                                In this dictionary also the species from the database SpeciesOrder are 
                                                added (other_species_dict) since these species are not found in the PEG
                                                database (The biovolume for these species can only be found on order 
                                                level).
                                                * {Keys: Order {Keys: Species, values: biovolume}}.
                                    
- 4. Order_avg_Biovolume_dict:                  For each order the average biovolume is based on the species values 
                                                inside this order.
                                                * { Keys: order, values: average biovolume per order}
                                    
- 5. Class_Species_avg_Biovolume:               Averaged biovolume per species nested in the corresponding taxanomy 
                                                group 'class'.
                                                * {Keys: Class {Keys: species, values: biovolume per species}}
                                                
- 6. Class_order_avg_Biovolume:                 Averaged biovolume per order nested in the corresponding taxanomy group
                                                'class'.
                                                * {Keys: Class {Keys: Order, values: biovolume per Order}}

- 7. Class_avg_Biovolume:                       Averaged biovolume per class based on the corresponding species.
                                                * {keys: Class, value: biovolume per class}
                                        
- 8. Group_Order_Species_avg_Biovolume_dict:    The functional groups (diatom, dionflagellates, flagellates and other) 
                                                are added to the different taxonomy order.
                                                * {Keys: group {keys: order, values: biovolume per species}}

'''
'----------------------------------------------------------------------------------------------------------------------'

species_Biovolume_dict = {}
''' 1. Creates a dictionary from the PEG database based on the species and the corresponding biovolume (µm3/L). This 
    dictionary will later be used to average the biovolume per species.
    
    - Defines the variable searchFor (the species name) and the corresponding data row (row_info).
    - If the unit of the measured biovolume is 'cell' the biovolume is added to the corresponding species.
    - Since there are multiple biovolume values for one species, all other values are appended to the species avoid
      doubling.
    - {Keys: Species, values: biovolume}.
    
    '''
for index, row in dfBiovolume.iterrows():
    searchFor = row['Species']
    row_info = row.values
    if row_info[16] == 'cell':
        if searchFor not in species_Biovolume_dict:  # add the species to the dictionary
            species_Biovolume_dict[searchFor] = [row_info[25]]  # row info column 25 is the species value for biovolume
            # in µm3/L
        else:  # if the species name is already written in the dictionary, add the found biovolume values
            species_Biovolume_dict[searchFor].append(row_info[25])

'----------------------------------------------------------------------------------------------------------------------'

species_avg_Biovolume_dict = {}
''' 2. Creates a dictionary in which the biovolume is averaged per species. This dictionary will be later nested in the 
    taxonomy group order and class level.
    
    - The species to be searchFor is checked on its biovolume value in the dictionary. The average biovolume value per 
      species is added after which divided by the amount of values, rounded by 4 decimals.
    - {Keys: Species, values: average biovolume per species}.
     
    '''
for searchFor, values_list in species_Biovolume_dict.items():  # goes trough the species_BV_dict made in previous step
    avg_BV = round(sum(values_list) / len(values_list), 4)
    species_avg_Biovolume_dict[searchFor] = avg_BV

'----------------------------------------------------------------------------------------------------------------------'

Order_Species_avg_Biovolume_dict = {}
''' 3. Creates a dictionary in which the averaged biovolume per species is nested in the taxonomy group 'order'. This 
    dictionary will later be used to calculate the average biovolume per order. This dict is created  since not all 
    species are found in the database at species level
    
    - Species and order are zipped to be able to call the variables spec and ord. Since the dictionary is empty the 
      order values are added as key with the species biovolume (average) dict as value. all other species values for 
      the order are added as additional value.
    - Some species are not singular cells (form filaments, colonies or coenobium). These species will not have a biovol
      value represented by 0 in the dictionary. Species that form filaments, colonies or coenobium is not covered in the
      scope of this script.
    - Species with their corresponding order from the Species Order database(other_species_dict are added to this 
      dicitonary and nested if the order is already found or added as a new order.
    - {Keys: Order {Keys: Species, values: biovolume}}.
    
    '''
Order_Species_avg_Biovolume_dict_unadapted = {}

for spec_name, ord_name in zip(Species, Order):
    if ord_name not in Order_Species_avg_Biovolume_dict_unadapted:
        Order_Species_avg_Biovolume_dict_unadapted[ord_name] = {spec_name: species_avg_Biovolume_dict.get(spec_name,
                                                                                                          int(0))}
    else:
        Order_Species_avg_Biovolume_dict_unadapted[ord_name][spec_name] = species_avg_Biovolume_dict.get(spec_name,
                                                                                                         int(0))

other_species_dict = dict(zip(Species_other, Species_other_order))  # creates a dictionary based on the Species_order
# database.

for spec, order_name in other_species_dict.items():
    if order_name in Order_Species_avg_Biovolume_dict_unadapted:
        Order_Species_avg_Biovolume_dict_unadapted[order_name][spec] = other_species_dict.get(order_name, 0)
    else:
        Order_Species_avg_Biovolume_dict_unadapted[order_name] = {spec: other_species_dict.get(order_name, 0)}

for order, biovolume in Order_Species_avg_Biovolume_dict_unadapted.items():
    order_key = str(order)
    if order_key == 'nan':
        order_key = 'Undefined order'
    elif order_key == ' ':
        order_key = 'Undefined order'
    else:
        order_key = order
    Order_Species_avg_Biovolume_dict[order_key] = biovolume

# pprint.pprint(Order_Species_avg_Biovolume_dict)


'----------------------------------------------------------------------------------------------------------------------'

Order_avg_Biovolume_dict = {}
''' 4. This dictionary creates an average biovolume for each order, based on the nested species biovolume values.

    - the keys order and species are called from the previous made dictionary.
    - A list is made for the biovolume values of the species biovolume values.
    - The average biovolume value per order is added and divided by the amount of values, rounded by 4 decimals.
    - {Keys: order, values: average biovolume per order}
    
    '''
for order, species in Order_Species_avg_Biovolume_dict.items():
    values_list = species.values()  # create a list of biovolume values for all species in the order
    avg_ord_BV = round(sum(values_list) / len(values_list), 4)  # calculate the average biovolume for the order
    Order_avg_Biovolume_dict[order] = avg_ord_BV  # store the average biovolume in the new dictionary

# pprint.pprint(Order_avg_Biovolume_dict)

'----------------------------------------------------------------------------------------------------------------------'

Class_Species_avg_Biovolume = {}
''' 5. Creates a dictionary in which the biovolume per corresponding species is nested in the taxonomy group 'class'. 

    - From the order species average biovolume dictionary the order and species are called after which the biovolume 
      value for the species could be defined.
    - All species with their corresponding classes are zipped in a new dictionary (SpeciesClass) to define the class per
      species.
    - If species in the dictionaries are equal the corresponding class is added to a new dictionary. The biovolume of 
      the species is added after which the new dictionary is added to the class species dictionary.
    - {Keys: Class {Keys: species, values: biovolume per species}}  

    '''
SpeciesClass = dict(zip(Species, Classes))  # species on class level

for order, spec in Order_Species_avg_Biovolume_dict.items():
    for sp, value in spec.items():
        for species, Cl in SpeciesClass.items():
            if species == sp:
                class_dict = Class_Species_avg_Biovolume.get(Cl, {})
                class_dict[sp] = value
                Class_Species_avg_Biovolume[Cl] = class_dict

# pprint.pprint(Class_Species_avg_Biovolume)

'----------------------------------------------------------------------------------------------------------------------'
Class_Order_dict = {}
''' 6. Dictionary in which the average order biovolume is nested in taxonomy group 'Class'. 

    - A new dictionary is made from the Species Order database in which the order is attached to the corresponding
      classes (OrderClass)
    - Since not all orders are defined in the Species Order database a try except approach is used to avoid key errors.
    - Classes are defined based on orders of the Order averaged biovolume in OrderClass.
    - classes, orders and the corresponding biovolume are then appended to the Class order dictionary.
      * {Keys: Class {Keys: order, values: biovolume}}.
    
    '''

OrderClass = dict(zip(Order, Classes))

for order, biovolume in Order_avg_Biovolume_dict.items():
    try:
        for ord_name in OrderClass:
            classes = OrderClass[order]
            if classes in Class_Order_dict:
                Class_Order_dict[classes][order] = biovolume
            else:
                Class_Order_dict[classes] = {order: biovolume}
    except KeyError:
        continue

# pprint.pprint(Class_Order_dict)

'----------------------------------------------------------------------------------------------------------------------'
Class_avg_Biovolume = {}
''' 7. Creates a dictionary in which the biovolume is averaged per class.

    - A new list is made from the biovolume values of the order from the class order dictionary, after which the average
      is calculated and rounded by 4 decimals.
    - The average biovolume is added per class. 
    - {keys: Class, value: biovolume per class}

    '''
for clss, orders in Class_Order_dict.items():
    biovolumes = []
    for values in orders.values():
        biovolumes.append(values)
    avg_biovolume = round(sum(biovolumes) / len(biovolumes), 4)
    Class_avg_Biovolume[clss] = avg_biovolume

# pprint.pprint(Class_avg_Biovolume)
'----------------------------------------------------------------------------------------------------------------------'

Group_Order_Species_avg_Biovolume_dict = {}
''' 8. Creates a dictionary for functional groups (diatoms, dinoflagellates, flagellates, other) with the corresponding 
    order and their species nested in it. This dictionary will provide an addition in conversion to order level and in 
    later stages to group the data in functional groups.
    
    - a new dictionary is made from the Species Order database, based on the order with their corresponding classes 
      (GroupOrder), so that all the species inside these orders can be appended to the Group Order Dictionary.
    - If orders from the GroupOrder and the Order averaged biovolume dictionary are equal the group and the order with 
      the corresponding biovolume are added. 
    - In the second step the species corresponding to the order are added. 
    - {Keys: group {keys: order, values: biovolume per species}}
    '''
GroupOrder = dict(zip(Species_other_order, Species_other_group))

for order, biovolume in Order_avg_Biovolume_dict.items():
    for orderdiff, group in GroupOrder.items():
        if orderdiff == order:
            if group in Group_Order_Species_avg_Biovolume_dict:
                Group_Order_Species_avg_Biovolume_dict[group][order] = biovolume
            else:
                Group_Order_Species_avg_Biovolume_dict[group] = {order: biovolume}

# Nest the species corresponding to the order in the SpOr_merge_BV_orderlevel_Group dictionary
for group, order in Group_Order_Species_avg_Biovolume_dict.items():
    for ord_name, spec in Order_Species_avg_Biovolume_dict.items():
        if ord_name in order:
            Group_Order_Species_avg_Biovolume_dict[group][ord_name] = spec

# pprint.pprint(Group_Order_Species_avg_Biovolume_dict)
'----------------------------------------------------------------------------------------------------------------------'
''' (To avoid scrolling back up) Overview of all dictonaries and their functions:

- 1. species_Biovolume_dict:                    Species from the PEG database with their corresponding biovolume values 
                                                (multiple values per species). 
                                                * {Keys: Species, values: biovolume}.
                                     
- 2. species_avg_Biovolume_dict:                Average biovolume for species from the PEG database. 
                                                * {Keys: Species, values: average biovolume per species}.

- 3. Order_Species_avg_Biovolume_dict:          The average biovolume per species is nested in the corresponding order. 
                                                In this dictionary also the species from the database SpeciesOrder are 
                                                added (other_species_dict) since these species are not found in the PEG
                                                database (The biovolume for these species can only be found on order 
                                                level).
                                                * {Keys: Order {Keys: Species, values: biovolume}}.
                                    
- 4. Order_avg_Biovolume_dict:                  For each order the average biovolume is based on the species values 
                                                inside this order.
                                                * { Keys: order, values: average biovolume per order}
                                    
- 5. Class_Species_avg_Biovolume:               Averaged biovolume per species nested in the corresponding taxanomy 
                                                group 'class'.
                                                * {Keys: Class {Keys: species, values: biovolume per species}}
                                                
- 6. Class_order_avg_Biovolume:                 Averaged biovolume per order nested in the corresponding taxanomy group
                                                'class'.
                                                * {Keys: Class {Keys: Order, values: biovolume per Order}}

- 7. Class_avg_Biovolume:                       Averaged biovolume per class based on the corresponding species.
                                                * {keys: Class, value: biovolume per class}
                                        
- 8. Group_Order_Species_avg_Biovolume_dict:    The functional groups (diatom, dionflagellates, flagellates and other) 
                                                are added to the different taxonomy order.
                                                * {Keys: group {keys: order, values: biovolume per species}}

'''
'----------------------------------------------------------------------------------------------------------------------'


def BiovolumeSpeciesLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    sample_id = []
    species_bsl = []
    result_spec = []
    concentration = []
    group_bsl = []
    for index, row in inputdata_dataframe.iterrows():
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        if searchFor in species_avg_Biovolume_dict:
            for group, order in Group_Order_Species_avg_Biovolume_dict.items():
                for ord, spec in order.items():
                    if searchFor in spec:
                        BV = round(species_avg_Biovolume_dict[searchFor] * Conc / 1e+9, 5)
                        if searchFor not in species_bsl:
                            sample_id.append(ID)
                            species_bsl.append(searchFor)
                            concentration.append(round(Conc, 2))
                            print(f"{searchFor} found in PEG database. The biovolume for {searchFor} is {BV} ml")
                            result_spec.append(BV)
                            group_bsl.append(group)
        else:
            print(f"{searchFor} not found in PEG database at any taxonomy level")
    if len(result_spec) > 0.0:
        df = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bsl,
            'Group': group_bsl,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_spec
        })
        return df


df = BiovolumeSpeciesLevel(  # define the column name and data of the dataset (InputData) to be converted below:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output.
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)


#
# '----------------------------------------------------------------------------------------------------------------------'
# #
def BiovolumeOrderLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    sample_id = []
    species_bol = []
    result_ord = []
    concentration = []
    group_bol = []
    for index, row in inputdata_dataframe.iterrows():
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        if searchFor not in species_avg_Biovolume_dict:
            for ord, species_dict in Order_Species_avg_Biovolume_dict.items():
                if searchFor in species_dict:
                    searchFor_order = ord
                    BV_order = round(Order_avg_Biovolume_dict[searchFor_order] * Conc / 1e+9, 5)
                    # print(f"Biovolume of {searchFor} found at order level ({ord}): {BV_order} ml")
                    if BV_order > 0.0:
                        sample_id.append(ID)
                        species_bol.append(searchFor)
                        concentration.append(round(Conc, 2))
                        result_ord.append(BV_order)
                        for ord in Group_Order_Species_avg_Biovolume_dict.values():
                            if searchFor_order in ord:
                                group_bol.append(group)
                    elif BV_order == 0.0:
                        continue
                        # print(f"See for biovolume for species {searchFor} class level")
                    break
                    # else:
                    #      print(f"Species {searchFor} biovolume not found at species and order level")
    if len(result_ord) > 0:
        df2 = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bol,
            'Group': group_bol,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_ord,
        })
        return df2


df2 = BiovolumeOrderLevel(  # define the column name and data of the dataset (InputData) to be converted below:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output.
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)
'----------------------------------------------------------------------------------------------------------------------'


def BiovolumeClassLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    sample_id = []
    species_bcl = []
    concentration = []
    result_class = []
    group_bcl = []
    for index, row in inputdata_dataframe.iterrows():
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        for ord, species_dict in Order_Species_avg_Biovolume_dict.items():
            for spec in species_dict.keys():
                if searchFor == spec:
                    searchFor_order = ord
                    BV_order = round(Order_avg_Biovolume_dict[searchFor_order] * Conc / 1e+9, 5)
                    if BV_order == 0:
                        for cls in Class_avg_Biovolume:
                            sample_id.append(ID)
                            concentration.append(round(Conc, 2))
                            species_bcl.append(searchFor)
                            searchFor_Class = cls
                            BV_Class = round(Class_avg_Biovolume[searchFor_Class] * Conc / 1e+9, 5)
                            # print(f"Biovolume of {searchFor} found at class level ({cls}): {BV_Class} ml")
                            result_class.append(BV_Class)
                            for group, Or in Group_Order_Species_avg_Biovolume_dict.items():
                                for o in Or.keys():
                                    if o == ord:
                                        group_bcl.append(group)
                            break
                        break
                    break
                break
                # else:
                #      print(f"Species {searchFor} biovolume not found at species and order level")
    if len(result_class) > 0:
        df3 = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bcl,
            'Group': group_bcl,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_class
        })
        return df3


df3 = BiovolumeClassLevel(  # define the column name and data of the dataset (InputData) to be converted below:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output.
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)

if df is not None:
    print(df)
elif df2 is not None:
    print(df2)
elif df3 is not None:
    print(df3)
else:
    print('No data found')
'----------------------------------------------------------------------------------------------------------------------'

#  creates an excel file with the data found at species (df), order (df2) and class level (df3).

with pd.ExcelWriter('Biovolume_test.xlsx') as writer:
    df.to_excel(writer,
                sheet_name='Biovolume Species',
                index=False)
    df2.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + 1,
                 index=False,
                 header=False)
    df3.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + len(df2) + 1,
                 index=False,
                 header=False)
#
'----------------------------------------------------------------------------------------------------------------------'

# function outputs to df (choose own path and file name!):
Output_dataframe = pd.read_excel('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx',
                                 sheet_name='Biovolume Species')


def BiovolumeToCarbon(biovolume, cells):
    carbon = []
    for index, row in Output_dataframe.iterrows():
        carbon_diatoms = (0.228 * row[biovolume] ** 0.811) * row[cells]
        carbon_other = (0.216 * row[biovolume] ** 0.939) * row[cells]
        Group = row['Group']
        if Group == 'Diatom':
            carbon_output = round(carbon_diatoms, 2)
            carbon.append(carbon_output)
        else:
            carbon_output = round(carbon_other, 2)
            carbon.append(carbon_output)
    if len(carbon) > 0:
        df4 = pd.DataFrame({
            'Carbon (pg C)': carbon
        })
        return df4


df4 = BiovolumeToCarbon(
    biovolume='Biovolume (ml)',
    cells='Concentration (cells L-1)')

'----------------------------------------------------------------------------------------------------------------------'
with pd.ExcelWriter('Biovolume_test.xlsx') as writer:
    df.to_excel(writer,
                sheet_name='Biovolume Species',
                index=False)
    df2.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + 1,
                 index=False,
                 header=False)
    df3.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + len(df2) + 1,
                 index=False,
                 header=False)
    df4.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startcol=5,
                 index=False)

Output_dataframe_Carbon = pd.read_excel('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx',
                                        sheet_name='Biovolume Species')

'----------------------------------------------------------------------------------------------------------------------'


def Sum(Concentration, Biovolume, Carbon):
    conc = []
    biovol = []
    carb = []
    sum_conc = round(sum(Concentration), 2)
    conc.append(sum_conc)
    sum_biovol = round(sum(Biovolume), 2)
    biovol.append(sum_biovol)
    sum_carb = round(sum(Carbon), 2)
    carb.append(sum_carb)
    df5 = pd.DataFrame({
        'Sum_Concentration': conc,
        'Sum_Biovolume': biovol,
        'Sum_Carbon': carb
    })
    return df5


df5 = Sum(
    Concentration=Output_dataframe_Carbon['Concentration (cells L-1)'],
    Biovolume=Output_dataframe_Carbon['Biovolume (ml)'],
    Carbon=Output_dataframe_Carbon['Carbon (pg C)'])

'----------------------------------------------------------------------------------------------------------------------'
Output_dataframe_Carbon = pd.read_excel('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx',
                                        sheet_name='Biovolume Species')


# todo: check function variables below

def AveragesFunctionGroups(Concentration, Biovolume, carbon):
    diatom_values = []
    flagellates_values = []
    dinoflagellates_values = []
    other_values = []
    sum_carbon_diatom = []
    sum_carbon_flag = []
    sum_carbon_dino = []
    sum_other = []
    for index, row in Output_dataframe_Carbon.iterrows():
        carbon = row['Carbon (pg C)']
        group = row['Group']
        if group == 'Diatom':
            diatom_values.append(carbon)
            sum_carbon_diatom = round(sum(diatom_values), 0)
        elif group == 'Flagellates':
            flagellates_values.append(carbon)
            sum_carbon_flag = round(sum(flagellates_values), 1)
        elif group == 'Dinoflagellates':
            dinoflagellates_values.append(carbon)
            sum_carbon_dino = round(sum(dinoflagellates_values), 1)
        else:
            other_values.append(carbon)
            sum_other = round(sum(other_values), 1)
    df6 = pd.DataFrame({
        'Carbon Diatoms (pg C)': sum_carbon_diatom,
        'Carbon Flagellates (pg C)': sum_carbon_flag,
        'Carbon Dinoflagellates (pg C)': sum_carbon_dino,
        'Carbon Other (pg C)': sum_other},
        index=[0])
    return df6


df6 = AveragesFunctionGroups(
    Concentration=Output_dataframe_Carbon['Concentration (cells L-1)'],
    Biovolume=Output_dataframe_Carbon['Biovolume (ml)'],
    carbon=Output_dataframe_Carbon['Carbon (pg C)'])

with pd.ExcelWriter('Output_data_conversion_to_carbon.xlsx') as writer:
    df.to_excel(writer,
                sheet_name='Biovolume Species',
                index=False)
    df2.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + 1,
                 index=False,
                 header=False)
    df3.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + len(df2) + 1,
                 index=False,
                 header=False)
    df4.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startcol=5,
                 index=False)
    df5.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startrow=len(df) + len(df2) + len(df3) + 1,
                 startcol=3,
                 index=False,
                 header=False)
    df6.to_excel(writer,
                 sheet_name='Biovolume Species',
                 startcol=7,
                 index=True,
                 header=True)

# Make workbook look nicer, create averages and graphs

Final_Output = openpyxl.load_workbook(
    'C:/GIS_Course/EGM722/Project/Project_script/Output_data_conversion_to_carbon.xlsx')
sheet = Final_Output['Biovolume Species']
data_range = sheet['A1': get_column_letter(sheet.max_column) + str(sheet.max_row)]

# freeze headers and adjust column width and sort data
sheet.freeze_panes = 'A2'

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 33
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 25
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['I'].width = 20
sheet.column_dimensions['J'].width = 23
sheet.column_dimensions['K'].width = 27
sheet.column_dimensions['L'].width = 20

# changing font
fontobj1 = Font(name='Calibri', size=12, bold=True, italic=False)
sheet['C' + str(sheet.max_row)] = 'Sum:'
sheet['C' + str(sheet.max_row)].font = fontobj1
sheet['D' + str(sheet.max_row)].font = fontobj1
sheet['E' + str(sheet.max_row)].font = fontobj1
sheet['F' + str(sheet.max_row)].font = fontobj1

# make averages for concentration, biovolume and carbon

refObj = Reference(sheet, min_col=9, max_col=12, min_row=1, max_row=2)

chartObj = BarChart()
chartObj.add_data(refObj, titles_from_data=True)
chartObj.title = 'Sum carbon functional groups'
chartObj.x_axis.title = 'Functional groups'
chartObj.y_axis.title = 'Carbon (pg C)'

sheet.add_chart(chartObj, 'I4')

Final_Output.save('Output_data_conversion_to_carbon.xlsx')

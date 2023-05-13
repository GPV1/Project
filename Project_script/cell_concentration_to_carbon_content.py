import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import os

# import datafiles and sheetname. Necessary: database files PEG, Species order and input data. Provide the correct path
# to the datafiles in the repository.
dfBiovolume = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/PEG_BVOL2019_PJ.xlsx', sheet_name='Biovolume file')
dfSpeciesOrder = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/Species_order_GPV.xlsx', sheet_name='Sheet1')
InputData = pd.read_excel('C:/GIS_Course/EGM722/Project/Data_files/Data_2018_GPV.xlsx', sheet_name='Sheet1')

# Defines the columns in the database for the taxonomy groups
Divisions = dfBiovolume['Division']
Classes = dfBiovolume['Class']
Order = dfBiovolume['Order'].str.title()  # Order names in PEG database are in capitals
Genus = dfBiovolume['Genus']
Species = dfBiovolume['Species']
Species_other = dfSpeciesOrder['spec_name'].str.replace('_', ' ')  # Species names in SpeciesOrder database have _
Species_other_order = dfSpeciesOrder['order']
Species_other_group = dfSpeciesOrder['group']

# To avoid Key errors and lower return of species in the output, certain signs or characters are removed.
InputData['spec_name'] = InputData['spec_name'].str.replace('.', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('<', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('>', '')
InputData['spec_name'] = InputData['spec_name'].str.replace('~', '')
InputDataAmount = len(InputData['spec_name'])

# to let the user know that the input data is being processed.
print(f"InputData being processed, the amount of species to search for: {InputDataAmount}")
print('......')

'----------------------------------------------------------------------------------------------------------------------'
''' In the following section the dictionaries are created. To create overview, already the function and content of all 
dictionaries is given:

- 1. species_Biovolume_dict:                    Species from the PEG database with their corresponding biovolume values 
                                                (multiple values per species). 
                                                * {Keys: Species, values: biovolume}.
                                     
- 2. species_avg_Biovolume_dict:                Average biovolume for species from the PEG database. 
                                                * {Keys: Species, values: average biovolume per species}.

- 3. Order_Species_avg_Biovolume_dict:          The average biovolume per species is nested in the corresponding order. 
                                                In this dictionary also the species from the database SpeciesOrder are 
                                                added (other_species_dict) since these species are not found in the PEG
                                                database (The biovolume for these species can only be found on order 
                                                level).
                                                * {Keys: Order {Keys: Species, values: biovolume}}.
                                    
- 4. Order_avg_Biovolume_dict:                  For each order the average biovolume is based on the species values 
                                                inside this order.
                                                * { Keys: order, values: average biovolume per order}
                                    
- 5. Class_Species_avg_Biovolume:               Averaged biovolume per species nested in the corresponding taxanomy 
                                                group 'class'.
                                                * {Keys: Class {Keys: species, values: biovolume per species}}
                                                
- 6. Class_order_avg_Biovolume:                 Averaged biovolume per order nested in the corresponding taxanomy group
                                                'class'.
                                                * {Keys: Class {Keys: Order, values: biovolume per Order}}

- 7. Class_avg_Biovolume:                       Averaged biovolume per class based on the corresponding species.
                                                * {keys: Class, value: biovolume per class}
                                        
- 8. Group_Order_Species_avg_Biovolume_dict:    The functional groups (diatom, dionflagellates, flagellates and other) 
                                                are added to the different taxonomy order.
                                                * {Keys: group {keys: order, values: biovolume per species}}

'''
'----------------------------------------------------------------------------------------------------------------------'

species_Biovolume_dict = {}
''' 1. Creates a dictionary from the PEG database based on the species and the corresponding biovolume (µm3/L). This 
    dictionary will later be used to average the biovolume per species.
    
    - Defines the variable searchFor (the species name) and the corresponding data row (row_info).
    - If the unit of the measured biovolume is 'cell' the biovolume is added to the corresponding species.
    - Since there are multiple biovolume values for one species, all other values are appended to the species avoid
      doubling.
    - {Keys: Species, values: biovolume}.
    
    '''
for index, row in dfBiovolume.iterrows():
    searchFor = row['Species']
    row_info = row.values
    if row_info[16] == 'cell':
        if searchFor not in species_Biovolume_dict:  # add the species to the dictionary
            species_Biovolume_dict[searchFor] = [row_info[25]]  # row info column 25 is the species value for biovolume
            # in µm3/L
        else:  # if the species name is already written in the dictionary, add the found biovolume values
            species_Biovolume_dict[searchFor].append(row_info[25])


'----------------------------------------------------------------------------------------------------------------------'

species_avg_Biovolume_dict = {}
''' 2. Creates a dictionary in which the biovolume is averaged per species. This dictionary will be later nested in the 
    taxonomy group order and class level.
    
    - The species to be searchFor is checked on its biovolume value in the dictionary. The average biovolume value per 
      species is added after which divided by the amount of values, rounded by 4 decimals.
    - {Keys: Species, values: average biovolume per species}.
     
    '''
for searchFor, values_list in species_Biovolume_dict.items():  # goes trough the species_BV_dict made in previous step
    avg_BV = round(sum(values_list) / len(values_list), 4)
    species_avg_Biovolume_dict[searchFor] = avg_BV  # appends the species name with the averaged biovolume

'----------------------------------------------------------------------------------------------------------------------'

Order_Species_avg_Biovolume_dict = {}
''' 3. Creates a dictionary in which the averaged biovolume per species is nested in the taxonomy group 'order'. This 
    dictionary will later be used to calculate the average biovolume per order. This dict is created  since not all 
    species are found in the database at species level
    
    - Species and order are zipped to be able to call the variables spec and ord. Since the dictionary is empty the 
      order values are added as key with the species biovolume (average) dict as value. all other species values for 
      the order are added as additional value.
    - Some species are not singular cells (form filaments, colonies or coenobium). These species will not have a biovol
      value represented by 0 in the dictionary. Species that form filaments, colonies or coenobium is not covered in the
      scope of this script.
    - Species with their corresponding order from the Species Order database(other_species_dict are added to this 
      dicitonary and nested if the order is already found or added as a new order.
    - {Keys: Order {Keys: Species, values: biovolume}}.
    
    '''
Order_Species_avg_Biovolume_dict_unadapted = {}  # this dictionary contains orders undefined and serves as a in between
                                                 # step

for spec_name, ord_name in zip(Species, Order):  # goes through Species and Order names
    if ord_name not in Order_Species_avg_Biovolume_dict_unadapted:  # append the order name if it is not in the dict and
        # add the species name and the corresponding biovolume from species average biovolume dict.
        Order_Species_avg_Biovolume_dict_unadapted[ord_name] = {spec_name: species_avg_Biovolume_dict.get(spec_name,
                                                                                                          int(0))}
        # if the species is not found in the species average biovolume dict a 0 is assigned as biovolume
    else:
        Order_Species_avg_Biovolume_dict_unadapted[ord_name][spec_name] = species_avg_Biovolume_dict.get(spec_name,
                                                                                                         int(0))
        # if the species name is already in the dictionary just add the corresponding biovolume.

other_species_dict = dict(zip(Species_other, Species_other_order))  # creates a dictionary based on the Species_order
# database.

for spec, order_name in other_species_dict.items():
    if order_name in Order_Species_avg_Biovolume_dict_unadapted:  # if the order is already in the dictionary
        Order_Species_avg_Biovolume_dict_unadapted[order_name][spec] = species_avg_Biovolume_dict.get(spec, 0)  # append
        # the species from the other species dictionary to the corresponding order in the order species average biovol
        # dictionary.
    else:
        Order_Species_avg_Biovolume_dict_unadapted[order_name] = {spec: species_avg_Biovolume_dict.get(spec, 0)}
        # if order not in dictionary add the order and nest the species names and retrieved biovolume with the matching
        # order.

for order, biovolume in Order_Species_avg_Biovolume_dict_unadapted.items():  # assign names to empty dictionary keys
    order_key = str(order)
    if order_key == 'nan':
        order_key = 'Undefined order'
    elif order_key == ' ':
        order_key = 'Undefined order'
    else:
        order_key = order
    Order_Species_avg_Biovolume_dict[order_key] = biovolume


'----------------------------------------------------------------------------------------------------------------------'

Order_avg_Biovolume_dict = {}
''' 4. This dictionary creates an average biovolume for each order, based on the nested species biovolume values.

    - the keys order and species are called from the previous made dictionary.
    - A list is made for the biovolume values of the species biovolume values.
    - The average biovolume value per order is added and divided by the amount of values, rounded by 4 decimals.
    - {Keys: order, values: average biovolume per order}
    
    '''
for order, species in Order_Species_avg_Biovolume_dict.items():
    values_list = species.values()  # create a list of biovolume values for all species in the order
    avg_ord_BV = round(sum(values_list) / len(values_list), 4)
    Order_avg_Biovolume_dict[order] = avg_ord_BV  # store the average biovolume in the new dictionary

'----------------------------------------------------------------------------------------------------------------------'

Class_Species_avg_Biovolume = {}
''' 5. Creates a dictionary in which the biovolume per corresponding species is nested in the taxonomy group 'class'. 

    - From the order species average biovolume dictionary the order and species are called after which the biovolume 
      value for the species could be defined.
    - All species with their corresponding classes are zipped in a new dictionary (SpeciesClass) to define the class per
      species.
    - If species in the dictionaries are equal the corresponding class is added to a new dictionary. The biovolume of 
      the species is added after which the new dictionary is added to the class species dictionary.
    - {Keys: Class {Keys: species, values: biovolume per species}}  

    '''
SpeciesClass = dict(zip(Species, Classes))  # dictionary for species on class level

for order, spec in Order_Species_avg_Biovolume_dict.items():
    for sp, value in spec.items():  # goes through the biovolume values of the species in the ord spec biovol dictionary
        for species, Cl in SpeciesClass.items():
            if species == sp:  # if the species in the species class dictionary is the same as the species in the ord
                # spec avg biovol dict
                class_dict = Class_Species_avg_Biovolume.get(Cl, {})  # append the class and create a empty dictionary
                # to add the species and biovolumes
                class_dict[sp] = value  # creates a dictionary in which the species is added as key and the biovolume as
                # value
                Class_Species_avg_Biovolume[Cl] = class_dict  # append the previous made dict to the corresponding class

# print(Class_Species_avg_Biovolume)

'----------------------------------------------------------------------------------------------------------------------'
Class_Order_dict = {}
''' 6. Dictionary in which the average order biovolume is nested in taxonomy group 'Class'. 

    - A new dictionary is made from the Species Order database in which the order is attached to the corresponding
      classes (OrderClass)
    - Since not all orders are defined in the Species Order database a try except approach is used to avoid key errors.
    - Classes are defined based on orders of the Order averaged biovolume in OrderClass.
    - classes, orders and the corresponding biovolume are then appended to the Class order dictionary.
      * {Keys: Class {Keys: order, values: biovolume}}.
    
    '''

OrderClass = dict(zip(Order, Classes)) # creates a dictionary for order and classes

for order, biovolume in Order_avg_Biovolume_dict.items():
    try:
        for ord_name in OrderClass:
            classes = OrderClass[order]  # retrieves the class corresponding to the order from the order class dict
            if classes in Class_Order_dict:  # if the class is already in the class order dictionary
                Class_Order_dict[classes][order] = biovolume  # add order as a key and the biovolume as value
            else:
                Class_Order_dict[classes] = {order: biovolume} # if not in the dict add a new key - value
    except KeyError:  # since in the order dictionary there are some undefined orders
        continue

'----------------------------------------------------------------------------------------------------------------------'
Class_avg_Biovolume = {}
''' 7. Creates a dictionary in which the biovolume is averaged per class.

    - A new list is made from the biovolume values of the species from the spec order dictionary, after which the average
      is calculated and rounded by 4 decimals.
    - The average biovolume is added per class. 
    - {keys: Class, value: biovolume per class}

    '''
for clss, spec in Class_Species_avg_Biovolume.items():
    biovolumes = []
    for values in spec.values():
        biovolumes.append(values)
    avg_biovolume = round(sum(biovolumes) / len(biovolumes), 4)
    Class_avg_Biovolume[clss] = avg_biovolume

# print(Class_avg_Biovolume)
'----------------------------------------------------------------------------------------------------------------------'

Group_Order_Species_avg_Biovolume_dict = {}
''' 8. Creates a dictionary for functional groups (diatoms, dinoflagellates, flagellates, other) with the corresponding 
    order and their species nested in it. This dictionary will provide an addition in conversion to order level and in 
    later stages to group the data in functional groups.
    
    - a new dictionary is made from the Species Order database, based on the order with their corresponding classes 
      (GroupOrder), so that all the species inside these orders can be appended to the Group Order Dictionary.
    - If orders from the GroupOrder and the Order averaged biovolume dictionary are equal the group and the order with 
      the corresponding biovolume are added. 
    - In the second step the species corresponding to the order are added. 
    - {Keys: group {keys: order, values: biovolume per species}}
    '''
GroupOrder = dict(zip(Species_other_order, Species_other_group))  # creates a dictionary of orders and groups from
# the Species_Order database.

for order, biovolume in Order_avg_Biovolume_dict.items():
    for orderdiff, group in GroupOrder.items():
        if orderdiff == order:  # if the order in the order_group dict is the same as the order in the order biovol dict
            if group in Group_Order_Species_avg_Biovolume_dict:  # if the group is already in the new dictionary
                Group_Order_Species_avg_Biovolume_dict[group][order] = biovolume  # append the group, order and the
                # biovolume
            else:
                Group_Order_Species_avg_Biovolume_dict[group] = {order: biovolume}  # if not in dictionary add the order
                # and biovolume as a new key-value entry.

# Nest the species corresponding to the order in the Group_Order_Species_avg_Biovolume dictionary
for group, order in Group_Order_Species_avg_Biovolume_dict.items():
    for ord_name, spec in Order_Species_avg_Biovolume_dict.items():
        if ord_name in order:  # if orders are equal add the species with their biovolumes to the dictionary
            Group_Order_Species_avg_Biovolume_dict[group][ord_name] = spec

# print(Group_Order_Species_avg_Biovolume_dict)
'----------------------------------------------------------------------------------------------------------------------'
''' (To avoid scrolling back up) Overview of all dictonaries and their functions:

- 1. species_Biovolume_dict:                    Species from the PEG database with their corresponding biovolume values 
                                                (multiple values per species). 
                                                * {Keys: Species, values: biovolume}.
                                     
- 2. species_avg_Biovolume_dict:                Average biovolume for species from the PEG database. 
                                                * {Keys: Species, values: average biovolume per species}.

- 3. Order_Species_avg_Biovolume_dict:          The average biovolume per species is nested in the corresponding order. 
                                                In this dictionary also the species from the database SpeciesOrder are 
                                                added (other_species_dict) since these species are not found in the PEG
                                                database (The biovolume for these species can only be found on order 
                                                level).
                                                * {Keys: Order {Keys: Species, values: biovolume}}.
                                    
- 4. Order_avg_Biovolume_dict:                  For each order the average biovolume is based on the species values 
                                                inside this order.
                                                * { Keys: order, values: average biovolume per order}
                                    
- 5. Class_Species_avg_Biovolume:               Averaged biovolume per species nested in the corresponding taxanomy 
                                                group 'class'.
                                                * {Keys: Class {Keys: species, values: biovolume per species}}
                                                
- 6. Class_order_avg_Biovolume:                 Averaged biovolume per order nested in the corresponding taxanomy group
                                                'class'.
                                                * {Keys: Class {Keys: Order, values: biovolume per Order}}

- 7. Class_avg_Biovolume:                       Averaged biovolume per class based on the corresponding species.
                                                * {keys: Class, value: biovolume per class}
                                        
- 8. Group_Order_Species_avg_Biovolume_dict:    The functional groups (diatom, dionflagellates, flagellates and other) 
                                                are added to the different taxonomy order.
                                                * {Keys: group {keys: order, values: biovolume per species}}

'''
'----------------------------------------------------------------------------------------------------------------------'


def BiovolumeSpeciesLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    '''
    This function converts the cell concentration (cells per liter) to a biovolume for each input species at taxonomy
    species level.

    function arguments:
    - ID_col_name: name for the column for the species sample ID or other ID that will be created in the output.
    - ID_data: The column in the input dataframe where the ID can be found.
    - Species_col_name: name for the column for the species name that will be created in the output.
    - Species_data: the column in the input dataframe where the species names can be found.
    - Conc_col_name: name for the column for the concentation of cells per liter that will be created in the output.
    - Conc_data: the column in the input dataframe where the concentration of cells per liter data can be found.

    A dataframe containing the input sample ID, species name, group, cell concentration and calculated biovolume is
    returned.

    '''
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    # Makes empty lists to append the found data to
    sample_id = []
    species_bsl = []
    result_spec = []
    concentration = []
    group_bsl = []  # bsl stands for biovolume species level
    for index, row in inputdata_dataframe.iterrows():
        # iterates over the ID, the species to search for and the cell concentration
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        if searchFor in species_avg_Biovolume_dict:  # look for the species at species level
            for group, order in Group_Order_Species_avg_Biovolume_dict.items():
                for ord, spec in order.items(): # loops through the species in the group order species dict
                    if searchFor in spec: # to later assign the right functional group
                        BV = round(species_avg_Biovolume_dict[searchFor] * Conc / 1e+9, 5) # biovolume * cell conc
                        # append all found data to the empty lists
                        sample_id.append(ID)
                        species_bsl.append(searchFor)
                        concentration.append(round(Conc, 2))
                        result_spec.append(BV)
                        group_bsl.append(group)
        else:  # if not found the species will be looked up at the next functions
            continue
    if len(result_spec) > 0:  # to avoid printing an empty dataframe in case something went wrong
        df = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bsl,
            'Group': group_bsl,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_spec
        })
        return df


df_biovol_species = BiovolumeSpeciesLevel( # define the column name and data of the dataset (InputData) to be converted:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output (string).
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)


#
# '----------------------------------------------------------------------------------------------------------------------'
# #
def BiovolumeOrderLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    '''
       This function converts the cell concentration (cells per liter) to a biovolume for each input species at taxonomy
       order level. It will check if there is not already a result returned at species level to avoid doubling.

       function arguments:
       - ID_col_name: name for the column for the species sample ID or other ID that will be created in the output.
       - ID_data: The column in the input dataframe where the ID can be found.
       - Species_col_name: name for the column for the species name that will be created in the output.
       - Species_data: the column in the input dataframe where the species names can be found.
       - Conc_col_name: name for the column for the concentation of cells per liter that will be created in the output.
       - Conc_data: the column in the input dataframe where the concentration of cells per liter data can be found.

       A dataframe containing the input sample ID, species name, group, cell concentration and calculated biovolume is
       returned.

       '''
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    # Makes empty lists to append the found data to
    sample_id = []
    species_bol = []
    result_ord = []
    concentration = []
    group_bol = []  # bol stands for biovolume order level
    for index, row in inputdata_dataframe.iterrows():
        # iterates over the ID, the species to search for and the cell concentration
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        if searchFor not in species_avg_Biovolume_dict:  # to avoid double output and only check not found species in
            # previous function
            for ord, species_dict in Order_Species_avg_Biovolume_dict.items():  # loop through the dict at order level
                if searchFor in species_dict:  # if species is found in the dictionary
                    searchFor_order = ord  # get the order and calculate the biovolume at order level
                    BV_order = round(Order_avg_Biovolume_dict[searchFor_order] * Conc / 1e+9, 5)
                    if BV_order > 0.0:  # since some orders do not have a biovolume, these will be searched for at class
                        # level
                        # append all found data to the empty lists
                        sample_id.append(ID)
                        species_bol.append(searchFor)
                        concentration.append(round(Conc, 2))
                        result_ord.append(BV_order)
                        for ord in Group_Order_Species_avg_Biovolume_dict.values():
                            if searchFor_order in ord:  # find the corresponding functional group for each species
                                group_bol.append(group)
                    elif BV_order == 0.0:  # ignore the orders that return 0
                        continue
                    break
    if len(result_ord) > 0:   # to avoid printing an empty dataframe in case something went wrong
        df2 = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bol,
            'Group': group_bol,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_ord,
        })
        return df2


df_biovol_order = BiovolumeOrderLevel(  # define the column name and data of the dataset (InputData) to be converted:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output.
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)
'----------------------------------------------------------------------------------------------------------------------'


def BiovolumeClassLevel(ID_col_name, ID_data, Species_col_name, Species_data, Conc_col_name, Conc_data):
    '''
        This function converts the cell concentration (cells per liter) to a biovolume for each input species at
        taxonomy class level.

        function arguments:
        - ID_col_name: name for the column for the species sample ID or other ID that will be created in the output.
        - ID_data: The column in the input dataframe where the ID can be found.
        - Species_col_name: name for the column for the species name that will be created in the output.
        - Species_data: the column in the input dataframe where the species names can be found.
        - Conc_col_name: name for the column for the concentation of cells per liter that will be created in the output.
        - Conc_data: the column in the input dataframe where the concentration of cells per liter data can be found.

        A dataframe containing the input sample ID, species name, group, cell concentration and calculated biovolume is
        returned.

    '''
    inputdata_dataframe = pd.DataFrame({ID_col_name: ID_data, Species_col_name: Species_data, Conc_col_name: Conc_data})
    # Makes empty lists to append the found data to
    sample_id = []
    species_bcl = []
    concentration = []
    result_class = []
    group_bcl = []  # stands for biovolume class level
    for index, row in inputdata_dataframe.iterrows():
        # iterates over the ID, the species to search for and the cell concentration
        ID = row[ID_col_name]
        searchFor = row[Species_col_name]
        Conc = row[Conc_col_name]
        for ord, species_dict in Order_Species_avg_Biovolume_dict.items():
            for spec in species_dict.keys():
                if searchFor == spec:
                    searchFor_order = ord
                    BV_order = round(Order_avg_Biovolume_dict[searchFor_order] * Conc / 1e+9, 5)
                    # calculates the biovolume at order level and checks if this was returned as 0
                    if BV_order == 0:
                        for cls in Class_avg_Biovolume:  # If so the biovolume will be calculated at class level
                            searchFor_Class = cls
                            BV_Class = round(Class_avg_Biovolume[searchFor_Class] * Conc / 1e+9, 5)
                            # appends all the found data to the empty lists
                            sample_id.append(ID)
                            concentration.append(round(Conc, 2))
                            species_bcl.append(searchFor)
                            result_class.append(BV_Class)
                            for group, Or in Group_Order_Species_avg_Biovolume_dict.items():
                                for o in Or.keys():  # find the corresponding functional group for the species
                                    if o == ord:
                                        group_bcl.append(group)
                            break
    if len(result_class) > 0:  # to avoid printing an empty dataframe in case something went wrong
        df3 = pd.DataFrame({
            ID_col_name: sample_id,
            Species_col_name: species_bcl,
            'Group': group_bcl,
            Conc_col_name: concentration,
            'Biovolume (ml)': result_class
        })
        return df3


df_biovol_class = BiovolumeClassLevel(  # define the column name and data of the dataset (InputData) to be converted:

    ID_col_name='ID',  # If there is an ID for the sample, name this column for in the output.
    ID_data=InputData['sample_ID'],  # define in the input data in which column (str(header)) the ID data can be found.
    Species_col_name='Species',  # name the column header for species data in the output.
    Species_data=InputData['spec_name'],  # define in the input data in which column the species data can be found.
    Conc_col_name='Concentration (cells L-1)',  # name the column header for concentration data in the output.
    Conc_data=InputData['conc_cells_per_L']  # define in the input data in which column the conc. data can be found.
)


# To show that the functions worked or to show that something went wrong
if df_biovol_species is not None:
    print('Match found in species dictionaries')
if df_biovol_order is not None:
    print('Match found in order dictionaries')
if df_biovol_class is not None:
    print('Match found in class dictionaries')
else:
    print('No match found in dictionaries, check input data')

'----------------------------------------------------------------------------------------------------------------------'

#  creates an Excel file with the data found in the functions for species (df_biovol_species), order (df_biovol_order)
#  and class level (df_biovol_class).

with pd.ExcelWriter('Biovolume_test.xlsx') as writer:
    df_biovol_species.to_excel(writer,
                                sheet_name='Biovolume Species',
                                index=False)
    df_biovol_order.to_excel(writer,
                                sheet_name='Biovolume Species',
                                startrow=len(df_biovol_species) + 1, # to write it below the first dataframe
                                index=False,
                                header=False)  # the first dataframe already has (the same) headers
    df_biovol_class.to_excel(writer,
                                sheet_name='Biovolume Species',
                                startrow=len(df_biovol_species) + len(df_biovol_order) + 1,  # to write it below the two df
                                index=False,
                                header=False)
#
'----------------------------------------------------------------------------------------------------------------------'

# To continue the script the combined data in the Excel file is written to a df (choose correct path and file name):
Output_dataframe = pd.read_excel('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx',
                                 sheet_name='Biovolume Species')

'----------------------------------------------------------------------------------------------------------------------'

def BiovolumeToCarbon(biovolume, cells):
    '''
        This function converts the calculated biovolume togther with the cell concentration to carbon concent in
        picogram carbon. Based on the species functional group one of the two equations is used (source: Menden-Deuer &
        Lessard (2000)):

        - diatoms = (0.228 * biovolume^0.811) * cell concentration
        - other = (0.216 * biovolume^0.939) * cell concentration

        Function arguments:
        - biovolume: the column header in the output dataframe in which the calculated biovolume can be found
        - cells: the column header in the output dataframe in which the cell concentration can be found

    '''
    # creates an empty list for the carbon results
    carbon = []
    for index, row in Output_dataframe.iterrows():
        carbon_diatoms = (0.228 * row[biovolume] ** 0.811) * row[cells]  # equation for diatoms
        carbon_other = (0.216 * row[biovolume] ** 0.939) * row[cells]  # equation for other functional groups
        Group = row['Group']  # to find the species functional group and calculate the carbon content
        if Group == 'Diatom':
            carbon_output = round(carbon_diatoms, 2)
            carbon.append(carbon_output)
        else:
            carbon_output = round(carbon_other, 2)
            carbon.append(carbon_output)
    if len(carbon) > 0:  # to avoid returning an empty dataframe
        df4 = pd.DataFrame({
            'Carbon (pg C)': carbon
        })
        return df4


df_carbon = BiovolumeToCarbon(
    biovolume='Biovolume (ml)',
    cells='Concentration (cells L-1)')

'----------------------------------------------------------------------------------------------------------------------'
# writes the carbon data in the previous created Excel file.
with pd.ExcelWriter('Biovolume_test.xlsx') as writer:
    df_biovol_species.to_excel(writer,
                                sheet_name='Biovolume Species',
                                index=False)
    df_biovol_order.to_excel(writer,
                                sheet_name='Biovolume Species',
                                startrow=len(df_biovol_species) + 1,
                                index=False,
                                header=False)
    df_biovol_class.to_excel(writer,
                                sheet_name='Biovolume Species',
                                startrow=len(df_biovol_species) + len(df_biovol_order) + 1,
                                index=False,
                                header=False)
    df_carbon.to_excel(writer,
                                sheet_name='Biovolume Species',
                                startcol=5,
                                index=False)


# to continue the script the output data including carbon is written to a dataframe (select the correct path):
Output_dataframe_Carbon = pd.read_excel('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx',
                                        sheet_name='Biovolume Species')

'----------------------------------------------------------------------------------------------------------------------'
def Sum(Concentration, Biovolume, Carbon):
    '''
    This function calculates the sum of the cell concentration, biovolume and carbon based on the output dataframe.

    Function arguments:
    - Concentration: the column in the output dataframe where the cell concentration can be found.
    - Biovolume: the column in the output dataframe where the biovolume can be found.
    - Carbon: The column in the output dataframe where the carbon content can be found.

    Returns a dataframe with the sum of the cell concentration, the biovolume and carbon content

    '''
    # Create empty lists to append result to
    conc = []
    biovol = []
    carb = []
    # Sum all cell concentration data, round it to two decimals and append to the list
    sum_conc = round(sum(Concentration), 2)
    conc.append(sum_conc)
    # Sum all biovolume data, round it to two decimals and append to the list
    sum_biovol = round(sum(Biovolume), 2)
    biovol.append(sum_biovol)
    # Sum all carbon content data, round it to two decimals and append to the list
    sum_carb = round(sum(Carbon), 2)
    carb.append(sum_carb)
    df5 = pd.DataFrame({
        'Sum_Concentration': conc,
        'Sum_Biovolume': biovol,
        'Sum_Carbon': carb
    })
    return df5


df_sum = Sum(
    Concentration=Output_dataframe_Carbon['Concentration (cells L-1)'],  # define the column in the output dataframe for
    # the cell concentration (output_dataframe_carbon['header']).
    Biovolume=Output_dataframe_Carbon['Biovolume (ml)'],  # and for the biovolume
    Carbon=Output_dataframe_Carbon['Carbon (pg C)'])  # and for the carbon content

'----------------------------------------------------------------------------------------------------------------------'
def SumFunctionGroups(carbon):
    '''
    This function calculates the sum of carbon per functional group (diatom, flagellates, dinoflagellates or others).

    Functional argument:
    - carbon: The column in the output dataframe where the carbon content can be found.

    Returns a data frame with the sum of carbon per functional group.
    '''
    # Creates lists where the found results can be appended to.
    diatom_values = []
    flagellates_values = []
    dinoflagellates_values = []
    other_values = []
    sum_carbon_diatom = []
    sum_carbon_flag = []
    sum_carbon_dino = []
    sum_other = []
    for index, row in Output_dataframe_Carbon.iterrows():  # iterates over the output dataframe and looks at the carbon
        # and the group of each species.
        carb = row[carbon]
        group = row['Group']
        # If diatom append the carbon value to the list and calculate the sum of this list
        if group == 'Diatom':
            diatom_values.append(carb)
            sum_carbon_diatom = round(sum(diatom_values), 0)
        # If flagellate append the carbon value to the list and calculate the sum of this list
        elif group == 'Flagellates':
            flagellates_values.append(carb)
            sum_carbon_flag = round(sum(flagellates_values), 1)
        # If dinoflagellate append the carbon value to the list and calculate the sum of this list
        elif group == 'Dinoflagellates':
            dinoflagellates_values.append(carb)
            sum_carbon_dino = round(sum(dinoflagellates_values), 1)
        # All others, append the carbon value to the list and calculate the sum of this list
        else:
            other_values.append(carb)
            sum_other = round(sum(other_values), 1)
    df6 = pd.DataFrame({
        'Carbon Diatoms (pg C)': sum_carbon_diatom,
        'Carbon Flagellates (pg C)': sum_carbon_flag,
        'Carbon Dinoflagellates (pg C)': sum_carbon_dino,
        'Carbon Other (pg C)': sum_other},
        index=['sum'])
    return df6


df_sum_functional_groups = SumFunctionGroups(carbon='Carbon (pg C)')  # define the column in the output dataframe for
# the carbon content ('header').

print('Carbon calculated')

'----------------------------------------------------------------------------------------------------------------------'
# All dataframes are written in the final Excel file
with pd.ExcelWriter('Output_data_conversion_to_carbon.xlsx') as writer:
    df_biovol_species.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        index=False)
    df_biovol_order.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        startrow=len(df_biovol_species) + 1,
                                        index=False,
                                        header=False)
    df_biovol_class.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        startrow=len(df_biovol_species) + len(df_biovol_order) + 1,
                                        index=False,
                                        header=False)
    df_carbon.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        startcol=5,
                                        index=False)
    df_sum.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        startrow=len(df_biovol_species)+len(df_biovol_order) + len(df_biovol_class) + 1,
                                        startcol=3,
                                        index=False,
                                        header=False)
    df_sum_functional_groups.to_excel(writer,
                                        sheet_name='Biovolume Species',
                                        startcol=7,
                                        index=True,
                                        header=True)


# Calculates the amount of species returned in the output data
outputData = len(Output_dataframe_Carbon['Species'])
# and the amount of species missing
missingData = InputDataAmount - outputData

'----------------------------------------------------------------------------------------------------------------------'
# Choose the right path for the final output Excel file.
Final_Output = openpyxl.load_workbook('C:/GIS_Course/EGM722/Project/Project_script/'
                                      'Output_data_conversion_to_carbon.xlsx')
# and select the right sheet.
sheet = Final_Output['Biovolume Species']

# freeze headers
sheet.freeze_panes = 'A2'  # freeze the first row

# change column width to the right dimensions
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 33
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 25
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['I'].width = 20
sheet.column_dimensions['J'].width = 23
sheet.column_dimensions['K'].width = 27
sheet.column_dimensions['L'].width = 20

# changing font for the sum of the biovolume, cell concentration and carbon content.
fontobj1 = Font(name='Calibri', size=12, bold=True, italic=False)
sheet['C' + str(sheet.max_row)] = 'Sum:'  # writes 'sum' in column C underneath the last data row
sheet['C' + str(sheet.max_row)].font = fontobj1  # sum in bold
sheet['D' + str(sheet.max_row)].font = fontobj1  # sum of cell concentration in bold
sheet['E' + str(sheet.max_row)].font = fontobj1  # sum of biovolume in bold
sheet['F' + str(sheet.max_row)].font = fontobj1  # sum of carbon content in bold

# Creating a bar chart for the sum of carbon content for the functional groups
refObj = Reference(sheet, min_col=9, max_col=12, min_row=1, max_row=2)  # selects the reference for the data
chartObj = BarChart()  # creates an empty bar chart
chartObj.add_data(refObj, titles_from_data=True)  # adds the data names as series names
chartObj.title = 'Sum carbon functional groups'  # creates a title for the bar chart
chartObj.x_axis.title = 'Functional groups'  # creates x axis title
chartObj.y_axis.title = 'Carbon (pg C)'  # creates y axis title

# adds the chart to the sheet
sheet.add_chart(chartObj, 'I4')

# saves the final output excel file and removes the biovolume file (no longer necessary, however if you want to keep it
# remove the remove function).
Final_Output.save('Output_data_conversion_to_carbon.xlsx')
os.remove('C:/GIS_Course/EGM722/Project/Project_script/Biovolume_test.xlsx')

'----------------------------------------------------------------------------------------------------------------------'
# to show that the script is finished and show the amount of species found, missing and refers to the Excel workbook to
# be opened for the results.
print('......')
print(f"InputData processing finished. From {InputDataAmount} entries, {outputData} are returned."
      f"\nNumber of species not converted to carbon: {missingData}."
      f"\nOpen Output_data_conversion_to_carbon.xlsx to see the results.")


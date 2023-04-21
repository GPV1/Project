import openpyxl
from openpyxl.styles import Font

'''' To open de data file and check the doc type'''
# open the data file
wb = openpyxl.load_workbook('Data_files/Data_2018.xlsx')

# Check the type of the excel file and get the sheet names
print(type(wb), wb.sheetnames)

''' To select the data to process and get the basic information'''
# select the right data sheet
data = wb['data_gesommeerd']

# Determine the size of the worksheet
print(data.max_row, data.max_column)

''' Check the basic commands and values in selected sheet'''
# Ways to access data in cells and get information on the position of the cell data
print(data['C179'].value)
same = data['C179']
print(same.value)
print(same.row)

print('Row {}, Column {} is {}'.format(same.row, same.column, same.value))
print('Cell {} is {}'. format(same.coordinate, same.value))

print(data.cell(row=1, column=2).value)

# To print a range of data from a certain column
for i in range(1,31,1):
    print(i, data.cell(row=i, column=3).value)

''' to print the data of each cell in a specific area '''
# first make a tuple of the data to display the its cell objects. Since there are 20 rows and 4 columns,
# rows gives a tuple of 20 tuples (each containing 4 cell objects),
# columns gives 4 tuples (each containing 20 cell objects)
tuple(data['A1':'D20'])

# two for loops are used. First one to goes over each row in the slice, 2nd loop goes through each cell in the loop
for rowOfCellObjects in data['A1':'D20']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
# to show the end of the tuple
    print('--- END OF ROW ---')

print(rowOfCellObjects)

# to access the values of cells in a particular row or column first convert to list
list(data.columns)[2]
for cellObj in list(data.columns)[2]:
    print(cellObj.value)

''' Create a new workbook '''
newwb = openpyxl.Workbook()
sheet = newwb.active
sheet.title = 'test'
print(newwb.sheetnames)

''' Save new workbook '''
newwb.save('test.xlsx')

''' Writing values to cells '''
testwb = openpyxl.load_workbook('test.xlsx')
print(testwb.sheetnames)

sheettest = testwb['test']
sheettest['A1'] = 'Check if this is written in the cell'
print(sheettest['A1'].value)

testwb.save('test.xlsx')

'''Setting the font style'''
fontobj1 = Font(name='Calibri', size=20, bold=True, italic=False)
sheettest['A1'].font = fontobj1

testwb.save('test.xlsx')

'''Using Formulas'''
sheettest['A2'] = 200
sheettest['A3'] = 300
sheettest['A4'] = '=SUM(A2:A3)'

testwb.save('test.xlsx')

'''Setting row height and column width'''
sheettest.row_dimensions[1].height = 30
sheettest.column_dimensions['A'].width = 60

testwb.save('test.xlsx')

'''Merging cells'''
sheettest.merge_cells('A1:C1')

testwb.save('test.xlsx')

# unmerge cells
sheettest.unmerge_cells('A1:C1')

testwb.save('test.xlsx')

'''Freezing Panes'''
sheettest.freeze_panes = 'A2'

testwb.save('test.xlsx')

'''Creating charts'''
# get some data in the test sheet
for i in range(1, 11):
    sheettest['C'+str(i)] = i

# create a reference in the sheet to put the graph in
refObj = openpyxl.chart.Reference(sheettest, min_col=3, min_row=1, max_col=3, max_row=10)

# create a series reference by passing in the reference object
seriesObj = openpyxl.chart.Series(refObj, title='First series')

# create a chart object
chartObj = openpyxl.chart.BarChart()
chartObj.tilte = 'Test chart'
chartObj.append(seriesObj)

# Add the chart to the sheet
sheettest.add_chart(chartObj, 'D1')

testwb.save('test.xlsx')


